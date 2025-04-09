from openpyxl.reader.excel import load_workbook
import pandas as pd
import os
import data.config as data
from config.log import log

class ReadExcel(object):

    @staticmethod
    def read_excel():
        if not os.path.exists(data.excel_path):
            print(f"文件路径错误或文件不存在: {data.excel_path}")
            return
        try:
            df = pd.read_excel(data.excel_path, engine='openpyxl')
            for index, row in df.iterrows():
                line = row
                yield line

        except Exception as e:
            print(f"读取文件时出错: {e}")


    @staticmethod
    def write_excel_to_cell(data, file_path=data.excel_path, sheet_name='Sheet1', startrow=0, startcol=0):
        """
        将数据写入指定单元格的Excel文件
        :param data: 要写入的数据，格式为列表的字典
        :param file_path: 写入的文件路径，默认为data.excel_path
        :param sheet_name: 写入的sheet名称，默认为'Sheet1'
        :param startrow: 写入的起始行，默认为0
        :param startcol: 写入的起始列，默认为0
        """
        try:
            # 检查文件是否存在
            if os.path.exists(file_path):
                # 加载现有工作簿
                workbook = load_workbook(file_path)
                if sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                else:
                    sheet = workbook.create_sheet(title=sheet_name)
            else:
                # 创建新的工作簿
                workbook = pd.ExcelWriter(file_path, engine='openpyxl')
                sheet = workbook.book.create_sheet(title=sheet_name)

            # 将数据写入指定单元格
            df = pd.DataFrame(data)
            for r_idx, row in df.iterrows():
                for c_idx, value in enumerate(row):
                    sheet.cell(row=startrow + r_idx + 1, column=startcol + c_idx + 1, value=value)

            # 保存工作簿
            workbook.save(file_path)
            log().info(f"数据已成功写入文件: {file_path}")
        except Exception as e:
            print(f"写入文件时出错: {e}")
